from openpyxl import load_workbook,Workbook
from datetime import date
from calendar import monthrange
from my_module.cellcontrol import Cell, Member
from copy import copy

# Variables
member_labels = ['編號','姓名','就職年','就職月','就職日','特休天數']
labels = ['早上班','早下班','晚上班','晚下班','早遲','早假','午遲','午假']
statistic_labels = ['工作天','加班1.33','加班1.67','遲到(分)','遲到(天)','特休(天)','病假(時)','事假(時)','其他(時)']

weekdays = ['Sun','Mon','Tue','Wed','Thr','Fri','Sat']
months = ['一月','二月','三月','四月','五月','六月','七月','八月','九月','十月','十一月','十二月']

'''Each .xlsx files contains records of whole year
each sheets in the .xlsx file contains records of a single month'''
try:
    today = date.today()
    book = load_workbook(f'{today.year}打卡紀錄.xlsx')
    #sys.stderr.write('檔案已存在，無法建置檔案。\n')
except FileNotFoundError:
    try:
        member_cell = Member('員工資料.xlsx')
        members = member_cell.members # members == [["a99","何宇智",2003,1,2,3], ["a98","何致泉",2007,7,2,7], ...]
        members.sort(key=lambda x:x[0]) # sort members
        
        ################################# 遞補員工編號區段
        # 新的一年，可以遞補員工編號
        member_numbers = [member[0] for member in members]
        new_numbers = []
        number = ['a',0]
        for i in member_numbers:
            if i[0] != number[0]:
                number = [i[0],0]
            number[-1] += 1
            new_numbers.append(f'{number[0]}{number[1]:02d}')
        for i,j in enumerate(new_numbers):
            members[i][0] = j
        ################################# 遞補員工編號區段
        
        new_member_cell = Cell(f'{today.year}員工資料.xlsx') # Create a new member file
        for label in member_labels:
            new_member_cell.write(data=label, align='center')
            new_member_cell.right()
        member_cell.to(2,1)
        new_member_cell.to(2,1)
        for n,i in enumerate(members):
            color = "grey" if n % 2 == 0 else False
            for j in i:
                member_cell.write(data=j, align="center", fill_color=color)
                new_member_cell.write(data=j, align="center", fill_color=color)
                member_cell.right()
                new_member_cell.right()
            member_cell.move(1,-6)
            new_member_cell.move(1,-6)
        member_cell.save(close=True)
        new_member_cell.save(close=True)
    except:
        #sys.stderr.write('查無員工資料，無法建置檔案。\n')
        assert False
    cell = Cell(f'{today.year}打卡紀錄.xlsx')
    cell.rename_sheet(f'{today.year}年度統計')
    cell.cell_size(width=4)
    cell.right()
    cell.cell_size(width=12)
    cell.right()
    cell.write('工作月', bold=True, align="center", fill_color="yellow")
    cell.cell_size(width=8)
    for i in statistic_labels:
        cell.right()
        cell.cell_size(width=10)
        cell.write(data=i, bold=True, align="center", fill_color="yellow")
    for label in ['特休天數', '已休天數', '剩餘天數']:
        cell.right()
        cell.write(data=label, bold=True, align="center", fill_color="bright_blue")
        cell.cell_size(width=11)
    cell.to(2,1)
    for n, member in enumerate(members):
        color = "grey" if n % 2 == 1 else False
        if member[2] <= today.year: # starts working from this year
            cell.write(data=member[0], bold=True, align="center", fill_color=color)
            cell.right()
            cell.write(data=member[1], bold=True, align="center", fill_color=color)
            cell.move(1,-1)
    for month in range(1,13):
        cell.create_sheet(f'{month:02d}月份資料')
        days = monthrange(today.year, month)[1]
        ######## set up all cell size (width) ########
        for i in [1,11,21]:
            cell.to(1,i)
            cell.cell_size(width=8)
        for i in [2,12,22]:
            cell.to(1,i)
            cell.cell_size(width=12)
        for i in [3,4,5,6,13,14,15,16,23,24,25,26]:
            cell.to(1,i)
            cell.cell_size(width=8)
        for i in [7,8,9,10,17,18,19,20,27,28,29,30]:
            cell.to(1,i)
            cell.cell_size(width=5.5)
        cell.to(1,32)
        cell.cell_size(width=4)
        cell.to(1,33)
        cell.cell_size(width=12)
        for i in range(34,43):
            cell.to(1,i)
            cell.cell_size(width=10)
        cell.right()
        cell.cell_size(width=15)
        ################
        cell.to(1,1)
        for n, member in enumerate(members):
            if member[2] > today.year or member[2] == today.year and member[3] > month:
                continue
            cell.write(data=member[0], bold=True, align="center", fill_color="bright_blue")
            cell.right()
            cell.write(data=member[1], bold=True, align="center", fill_color="bright_blue")
            for n, label in enumerate(labels):
                cell.right()
                if n % 2 == 0:
                    color = "yellow"
                else:
                    color = False
                cell.write(data=label, bold=True, align="center", fill_color=color)
            cell.move(1,-len(labels)-1)
            for day in range(days):
                cell.write(data=day+1, bold=True, align="center")
                cell.down()
            cell.move(-days-1,10)
            if cell.current_pos[1] == 31:
                cell.move(days+2,-30)
        cell.to(1,33)
        for n,label in enumerate(statistic_labels):
            cell.right()
            if n % 2 == 0:
                color = "yellow"
            else:
                color = False
            cell.write(data=label, bold=True, align="center", fill_color=color)
        cell.right()
        cell.write("特休剩餘天數", bold=True, align="center")
        cell.move(1,-len(statistic_labels)-2)
        for member in members:
            if member[2] > today.year or member[2] == today.year and member[3] > month:
                continue
            elif cell.current_pos[0] % 2 == 0:
                color = "bright_yellow"
            else:
                color = False
            cell.write(data=member[0], bold=True, align="center", fill_color=color)
            cell.right()
            cell.write(data=member[1], bold=True, align="center", fill_color=color)
            cell.move(1,-1)
    cell.save(close=True)
try:
    holiday_book = load_workbook(f'{today.year}假日表.xlsx')
except FileNotFoundError:
    cell = Cell(f'{today.year}假日表.xlsx')
    cell.rename_sheet(f'{today.year}假日表')
    for month in range(1,13):
        if month in [1,3,6,8,9,11]:
            color = "grey"
        else:
            color = None
        cell.set_position()
        cell.write(data=months[month-1], bold=True, align="center", fill_color=color)
        cell.cell_size(width=8)
        cell.right()
        for i in weekdays:
            cell.write(i, bold=True, align="center", fill_color=color)
            cell.cell_size(width=5)
            cell.right()
        cell.down()
        weekday, days = monthrange(today.year, month)
        cell.left((5-weekday)%7+1)
        for day in range(days):
            day += 1
            cell.write(data=day, align="center")
            cell.right()
            if cell.current_pos[1] % 8 == 1:
                cell.move(1,-7)
        cell.pin()
        if month % 4 == 0:
            cell.move(7,-24)
        else:
            cell.right(8)
    cell.save(close=True)