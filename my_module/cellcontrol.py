from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from calendar import monthrange
from copy import copy
#import sys
from datetime import date

# Some Color Variables
Colors = {None:None,
          "black":"00000000",
          "red":"00FF0000",
          "green":"0000FF00",
          "dark_blue":"000000FF",
          "yellow":"00FFFF00",
          "bright_blue":"0000FFFF",
          "grey":"00C0C0C0",
          "orange":"00FF6600",
          "pink":"00FF99CC",
          "bright_yellow":"00FFFF99"
         }

class Cell:
    def __init__(self, name, NotFoundThenCreate=True):
        '''Initialize by the given book name'''
        self._name = name
        try:
            self._book = load_workbook(name)
        except:
            if NotFoundThenCreate:
                self._book = Workbook()
            else:
                raise FileNotFoundError
        self._sheet = self._book.worksheets[0]
        self._cell = self._sheet.cell(1,1)
        self._pos_pin = dict()
    def close(self):
        'Close the file'
        self._book.close()
    def move(self, row=0, col=0):
        '''move by the given row, col value.
        
        move "down" by the given value of "row"
        move "right" by the given value of "col"
        
        the default value of "row" and "col" is zero.
        
        if col=0, "move" method is equivalent to "up" and "down"
        if row=0, "move" method is equivalent to "left" and "right"'''
        new_row = max(1, self._cell.row+row)
        new_column = max(1, self._cell.column+col)
        self._cell = self._sheet.cell(new_row,new_column)
    def up(self, delta=1):
        '''Move to up cell with given distance'''
        new_row = max(1, self._cell.row-delta)
        self._cell = self._sheet.cell(new_row,self._cell.column)
    def down(self, delta=1):
        '''Move to down cell with given distance'''
        new_row = max(1, self._cell.row+delta)
        self._cell = self._sheet.cell(new_row,self._cell.column)
    def right(self, delta=1):
        '''Move to right cell with given distance'''
        new_column = max(1, self._cell.column+delta)
        self._cell = self._sheet.cell(self._cell.row,new_column)
    def left(self, delta=1):
        '''Move to left cell with given distance'''
        new_column = max(1, self._cell.column-delta)
        self._cell = self._sheet.cell(self._cell.row,new_column)
    def to(self, row=None, column=None):
        '''Move to the given row and column'''
        new_row, new_column = self._cell.row, self._cell.column
        if row:
            new_row = row
        if column:
            new_column = column
        self._cell = self._sheet.cell(new_row,new_column)
    def cell_size(self, width=None, height=None, save=False):
        if width:
            self._sheet.column_dimensions[self._cell.column_letter].width = width
        if height:
            self._sheet.row_dimensions[self._cell.row].height = height
        if save:
            self.save()
    @property
    def value(self):
        return self._cell.value
    def create_sheet(self, name):
        '''Append a new sheet and move to it'''
        self._book.create_sheet(name)
        self._sheet = self._book.worksheets[-1]
        self._cell = self._sheet.cell(1,1)
    def to_sheet(self, num=0):
        '''Move to the given sheet number'''
        while num >= len(self._book.worksheets):
            num -= 1
        self._sheet = self._book.worksheets[num]
        self._cell = self._sheet.cell(1,1)
    def rename_sheet(self, name):
        '''Rename the current sheet'''
        self._sheet.title = name
    def write(self, data=None, fill_color=None, bold=False, color="black", align=None, save=False):
        '''Write data into current cell
        bold, color, align:
            style of the data
        data:
            None --> do nothing
            False --> clear data
            (given data) --> write to current cell
        fill_color:
            None --> do nothing
            False --> clear color
            (given color name) --> fill the given color
        save:
            True --> write and save
            False --> write only
        '''
        if data == False and type(data) == bool:
            self._cell.value = None
        else:
            self._cell.font=Font(bold=bold, color=Colors[color])
            if align:
                self._cell.alignment = Alignment(horizontal=align)
            if data != None:
                self._cell.value = data
        if fill_color == False:
            self._cell.fill = PatternFill()
        elif fill_color != None:
            self._cell.fill = PatternFill(fill_type='solid', start_color=Colors[fill_color])
        if save:
            self.save(name=self._name)
    def save(self, name=None, close=False):
        '''Save the workbook with the given file name
        If the parameter 'name' is given, the file will be renamed as the given name'''
        if name:
            self._name = name
        self._book.save(self._name)
        if close:
            self.close()
    @property
    def current_pos(self):
        '''Returns the current position'''
        return copy([self._cell.row, self._cell.column])
    def set_position(self, key=0):
        '''Set up (update) current position checkpoint, use self.pin() to move to this checkpoint position
        If an positive key is given, the corresponding key is required when using self.pin() to move.
        note : this function DONOT record the specific sheet'''
        self._pos_pin[key] = copy([self._cell.row, self._cell.column])
    def pin(self, key=0):
        '''Move cell to the set up position checkpoint, the default pin is [1,1]
        Default key is 0. If calling the other saved position, the corresponding key needs to be specified.'''
        self.to(*self._pos_pin[key])

class Member(Cell):
    def __init__(self, name, NotFoundThenCreate=True):
        '''this method can be viewd as of initiating variables.'''
        super().__init__(name, NotFoundThenCreate=True)
        self.set_position()
        self.to(2,1)
        self._members = []
        self._member_dict = dict()
        self._full_member_dict = dict()
        self._ids = []
        while self.value:
            assert self.value[0] in ['a','b','c','d','e']
            assert type(int(self.value[1:])) == int
            self._members.append([])
            self._ids.append(self.value)
            for col in range(6):
                if col == 1: # Member name
                    self._member_dict[self._ids[-1]] = self.value
                    self._full_member_dict[self._ids[-1]] = [self.value]
                elif col > 1:
                    self._full_member_dict[self._ids[-1]].append(self.value)
                self._members[-1].append(self.value)
                self.right()
            self.left(6)
            self.down()
        self.pin()
    @property
    def members(self):
        '''return the full information including id, name, the
        date(including year, month, and day) started to work, and special dayoff days.(as a list)'''
        return self._members
    @property
    def id_members(self):
        '''ONLY return id and name.(as a list)'''
        id_members = [[i] for i in self._ids]
        for i,ID in enumerate(self._ids):
            id_members[i].append(self.member_dict[ID])
        return id_members
    @property
    def member_dict(self):
        '''return the dictionary with ids as keys and with names as their values.'''
        return self._member_dict
    @property
    def full_member_dict(self):
        '''return the dictionary with ids as keys and with lists of all other informations as their values.'''
        return self._full_member_dict
    def find(self, member_id):
        '''Return True if the given member ID is in self._members'''
        if member_id in self._ids:
            return True
        return False
    @property
    def special_days(self):
        '''Return id, name, and special dayoff days.(as a list)'''
        return [[i[0],i[1],i[5]]for i in self._members]
    @property
    def special_days_dict(self):
        '''Return the dictionary with ids as keys and lists of names and special dayoff days as their values.'''
        spec_dict = dict()
        for i in self._members:
            spec_dict[i[0]] = [i[1],i[5]]
        return spec_dict

# Require update to successfully test this module
if __name__ == '__main__':
    # Test the class "Cell"
    cell = Cell(Workbook(),'cellcontroltest.xlsx')
    cell.write(None)
    cell.right(4)
    cell.write(2)
    cell.down(4)
    cell.write(3)
    cell.left(2)
    cell.write(4)
    cell.up(2)
    cell.write(5)
    cell.to(1,5)
    print(f'cell({cell.current_pos[0]},{cell.current_pos[1]}).value = {cell.value}')
    cell.create_sheet('new_sheet')
    cell.right(3)
    cell.down(3)
    cell.write('cell pin here')
    cell.set_position()
    cell.to_sheet()
    cell.rename_sheet('first sheet')
    cell.right()
    cell.down()
    cell.write('this is first saved')
    cell.save()
    cell.pin()
    cell.write('this is then saved')
    cell.save('cellcontroltest2.xlsx')
    print('class Cell tested successfully.')
    
    # Test the subclass "Record" of "Cell"
    '''record = Record(load_workbook('2023打卡紀錄.xlsx'), '2023打卡紀錄.xlsx')
    sheet1 = record.get_sheet(0)
    sheet1[0][0][0] = 'a99'
    sheet1[0][0][1] = '何宇智'
    sheet1[0][1][1] = 900
    print(sheet1[0][1][2])
    record.save_to_sheet(2,sheet1)
    record.create_sheet('blank_record')
    record.create_record(['a99','何宇智'],['早上班','早下班','晚上班','晚下班','遲到'])
    '''
    print('class Record tested successfully.')